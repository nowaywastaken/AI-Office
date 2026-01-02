from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from typing import Optional, List, Dict, Any, Union
from io import BytesIO


class ExcelEngine:
    """
    A comprehensive Excel workbook generator with full formatting control.
    Supports: cell formatting, formulas, charts, conditional formatting, and more.
    """
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
    
    # ==================== WORKSHEET OPERATIONS ====================
    
    def set_sheet_name(self, name: str):
        """Set the active worksheet name."""
        self.ws.title = name
        return self
    
    def add_sheet(self, name: str):
        """Add a new worksheet."""
        self.ws = self.wb.create_sheet(title=name)
        return self
    
    def select_sheet(self, name: str):
        """Select a worksheet by name."""
        self.ws = self.wb[name]
        return self
    
    # ==================== CELL OPERATIONS ====================
    
    def set_cell(
        self,
        row: int,
        col: int,
        value: Any,
        font_name: Optional[str] = None,
        font_size: Optional[int] = None,
        bold: bool = False,
        italic: bool = False,
        color: Optional[str] = None,
        bg_color: Optional[str] = None,
        alignment: str = 'left',
        number_format: Optional[str] = None
    ):
        """
        Set a cell value with formatting.
        
        Args:
            row, col: 1-indexed row and column
            value: Cell value
            font_name: Font family
            font_size: Font size
            bold, italic: Text decoration
            color: Font color (hex without #)
            bg_color: Background color (hex without #)
            alignment: 'left', 'center', 'right'
            number_format: Excel number format (e.g., '#,##0.00', '0%')
        """
        cell = self.ws.cell(row=row, column=col, value=value)
        
        # Font styling
        cell.font = Font(
            name=font_name or 'Arial',
            size=font_size or 11,
            bold=bold,
            italic=italic,
            color=color.lstrip('#') if color else None
        )
        
        # Alignment
        align_map = {'left': 'left', 'center': 'center', 'right': 'right'}
        cell.alignment = Alignment(horizontal=align_map.get(alignment, 'left'))
        
        # Background color
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color.lstrip('#'), 
                                    end_color=bg_color.lstrip('#'), 
                                    fill_type='solid')
        
        # Number format
        if number_format:
            cell.number_format = number_format
        
        return cell
    
    def set_formula(self, row: int, col: int, formula: str):
        """Set a formula in a cell."""
        self.ws.cell(row=row, column=col, value=formula)
        return self
    
    def set_row_data(
        self,
        row: int,
        data: List[Any],
        start_col: int = 1,
        header: bool = False
    ):
        """Set a row of data."""
        for i, value in enumerate(data):
            cell = self.ws.cell(row=row, column=start_col + i, value=value)
            if header:
                cell.font = Font(bold=True)
        return self
    
    def set_data_range(
        self,
        data: List[List[Any]],
        start_row: int = 1,
        start_col: int = 1,
        header: bool = True
    ):
        """Set a 2D range of data."""
        for i, row_data in enumerate(data):
            self.set_row_data(
                row=start_row + i,
                data=row_data,
                start_col=start_col,
                header=(header and i == 0)
            )
        return self
    
    # ==================== COLUMN/ROW SIZING ====================
    
    def set_column_width(self, col: int, width: float):
        """Set column width."""
        self.ws.column_dimensions[get_column_letter(col)].width = width
        return self
    
    def set_row_height(self, row: int, height: float):
        """Set row height."""
        self.ws.row_dimensions[row].height = height
        return self
    
    def auto_fit_columns(self, min_width: int = 10, max_width: int = 50):
        """Auto-fit column widths based on content."""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    continue  # Skip cells with non-stringifiable values
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            self.ws.column_dimensions[column_letter].width = adjusted_width
        return self
    
    # ==================== BORDERS ====================
    
    def add_borders(
        self,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        style: str = 'thin'
    ):
        """Add borders to a range of cells."""
        side = Side(style=style)
        border = Border(left=side, right=side, top=side, bottom=side)
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                self.ws.cell(row=row, column=col).border = border
        return self
    
    # ==================== CHARTS ====================
    
    def add_bar_chart(
        self,
        data_range: tuple,  # (min_col, min_row, max_col, max_row)
        position: str = 'E2',
        title: str = 'Chart'
    ):
        """Add a bar chart."""
        chart = BarChart()
        chart.title = title
        data = Reference(
            self.ws,
            min_col=data_range[0],
            min_row=data_range[1],
            max_col=data_range[2],
            max_row=data_range[3]
        )
        chart.add_data(data, titles_from_data=True)
        self.ws.add_chart(chart, position)
        return chart
    
    # ==================== SAVE OPERATIONS ====================
    
    def save(self, filepath: str):
        """Save workbook to file."""
        self.wb.save(filepath)
        return filepath
    
    def save_to_bytes(self) -> bytes:
        """Save workbook to bytes (for streaming)."""
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def create_blank(self):
        """Return the underlying workbook object."""
        return self.wb
